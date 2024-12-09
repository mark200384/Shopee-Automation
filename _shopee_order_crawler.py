import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import os

print("Current working directory:", os.getcwd())


# 初始化 Chrome 驅動器選項
def get_chrome_options():
    options = webdriver.ChromeOptions()
    current_dir = os.path.dirname(os.path.abspath(__file__))
    user_data_dir = os.path.join(current_dir, 'Chrome for Testing', 'User Data')
    print(f"Current directory: {current_dir}")  # 调试输出当前目录
    print(f"User data directory: {user_data_dir}")  # 调试输出用户数据目录
    options.add_argument(f'user-data-dir={user_data_dir}')
    options.add_argument('--profile-directory=Profile 1')
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--disable-gpu')
    return options


options = get_chrome_options()

# 使用 webdriver-manager 自動處理驅動器
from selenium.webdriver.chrome.service import Service

# service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(options=options)
driver.get('https://seller.shopee.tw/portal/sale/shipment?type=toship&source=all&sort_by=create_date_desc')
wait = WebDriverWait(driver, 10)


def transition_time(original_date):
    from datetime import datetime
    formatted_date = datetime.strptime(original_date, "%y%m%d").strftime("%Y/%m/%d")
    return formatted_date


def transition_delivery_way(text):
    if text == '蝦皮店到店':
        return '蝦皮'
    elif '7' in text:
        return '7-11'
    elif 'ok' in text.lower():
        return 'OK'
    return text


def filter_emoji(desstr, restr=''):
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(restr, desstr)

try:
    workbook = openpyxl.load_workbook('test2.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(title="order", index=0)


def write_order_xlsx(datas):
    # 將數據寫入一行
    for data in datas:
        sheet.append(data)

    current_row = sheet.max_row
    # print(current_row)
    for col in range(1, 7):
        sheet.merge_cells(start_row=max(current_row - len(datas) + 1, 1), start_column=col, end_row=current_row,
                          end_column=col)
    for col in range(15, 16):
        sheet.merge_cells(start_row=max(current_row - len(datas) + 1, 1), start_column=col, end_row=current_row,
                          end_column=col)
    sheet.merge_cells(start_row=max(1, current_row - len(datas) + 1), start_column=10, end_row=current_row,
                      end_column=10)
    for row in sheet.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=15):
        for cell in row:
            cell.font = Font(name=u'微軟正黑', size=10)
            cell.alignment = Alignment(vertical='center', horizontal='center')
    workbook.save('test2.xlsx')


# open_and_login(driver)

while True:
    try:
        # 嘗試找到元素
        order_list_table_shipment = wait.until(
            EC.presence_of_element_located((By.CLASS_NAME, 'order-list-table-shipment')))
        print("元素已找到，可以進行後續操作")
        break  # 如果成功找到元素，跳出迴圈
    except TimeoutException:
        print("請先登入")
        # 可以在這裡添加登入的程式碼或是等待用戶操作
        # 然後再次嘗試，或是根據實際情況進行適當的休眠
        time.sleep(5)  # 休眠5秒再次檢查，避免過於頻繁的請求

wrapper = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'table-body-wrapper')))
orders = wrapper.find_elements(By.CLASS_NAME, 'order-card')
t = 0
for order in reversed(orders):
    t = t + 1
    print("into loop i:{}".format(t))
    try:
        order_card_header = WebDriverWait(order, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'order-card-header'))
        )
        date = order.find_element(By.CLASS_NAME, 'order-sn').text[5:11]
        date = transition_time(date)
        buyer_id = order.find_element(By.CLASS_NAME, 'buyer-username').text
        name = ""
        delivery_way = order.find_element(By.CLASS_NAME, "fulfilment-channel-name").text
        delivery_way = transition_delivery_way(delivery_way)
        delivery_shop = ""
        phone = ""
        total_price = order.find_element(By.CLASS_NAME, "total-price").text
        total_price = int(total_price[3:].replace(',', ''))
        items = order.find_elements(By.CLASS_NAME, 'item-info')

        data = []
        for item in items:
            item_name = item.find_element(By.CLASS_NAME, 'item-name').text
            item_name = filter_emoji(item_name)
            item_description = item.find_element(By.CLASS_NAME, 'item-description').text[6:]  # 規格 ； 去除"商品規格: "
            color, size = "", ""
            if ',' in item_description:
                color, size = item_description.split(',')
            elif item_description.isnumeric():
                size = item_description
            else:
                color = item_description
            item_amount = item.find_element(By.CLASS_NAME, 'item-amount').text
            item_amount = int(item_amount[1:])  # remove "x"
            note = ""
            for _ in range(item_amount):
                data.append(
                    [date, buyer_id, name, phone, delivery_way, delivery_shop, item_name, color, size,
                     total_price, None, None, None, None, note])
        write_order_xlsx(data)
        # driver.close()
        # driver.switch_to.window(driver.window_handles[0])
    except NoSuchElementException:
        print('Element not found within the given time')
        pass
