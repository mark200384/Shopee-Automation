import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import re
import os
option = webdriver.ChromeOptions()
current_dir = os.path.abspath(os.path.dirname(__file__))
user_data_dir = os.path.join(current_dir, 'Chrome for Testing', 'User Data')
print(user_data_dir)
option.add_argument(fr'user-data-dir={user_data_dir}')
option.add_argument(r'--profile-directory=Default')
option.add_argument('--ignore-certificate-errors')
option.add_argument('--disable-gpu')
driver = webdriver.Chrome(ChromeDriverManager().install(), options = option)
driver.get('https://seller.shopee.tw/portal/notification/order-updates/')
wait = WebDriverWait(driver, 10)

def transition_time(date_str):
    from datetime import datetime
    date_object = datetime.strptime(date_str, "%Y/%m/%d %H:%M")

    # 將 datetime 物件格式化為新的日期字串
    formatted_date_str = date_object.strftime("%m/%d")
    return formatted_date_str

def transition_delivery_way(text):
    if text == '蝦皮店到店':
        return '蝦皮'
    elif text == '7-ELEVEN':
        return '7-11'
    elif text == 'OK MART':
        return 'OK'
    return text

def filter_emoji(desstr, restr=''):
    try:
        co = re.compile(u'[\U00010000-\U0010ffff]')
    except re.error:
        co = re.compile(u'[\uD800-\uDBFF][\uDC00-\uDFFF]')
    return co.sub(restr, desstr)

def process_product_name(text):
    product_name=''
    color=''
    size=''
    text = filter_emoji(text)
    if '較長備貨' in text:
        text = text[4:]
    if "規格:" in text:
        match = re.search(r'(.+?)\n規格: (.+)', text)
        if match:
            product_name = match.group(1)
            specifications = match.group(2)
            if ',' in specifications:
                color, size = specifications.split(',')
            else:
                color = specifications
    else:
        product_name = text
    return product_name, color, size


try:
    workbook = openpyxl.load_workbook('test2.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(title="order", index=0)
def write_order_xlsx(datas):
    # 將數據寫入一行
    for data in datas:
        sheet.append(data)

    current_row = sheet.max_row
    # print(current_row)
    for i in range(1,7):
        sheet.merge_cells(start_row=max(current_row - len(datas)+1,1), start_column=i, end_row=current_row, end_column=i)

    sheet.merge_cells(start_row=max(1,current_row - len(datas)+1), start_column=10, end_row=current_row, end_column=10)
    for row in sheet.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=10):
        for cell in row:
            cell.font = Font(name=u'微軟正黑', size=10)
            cell.alignment = Alignment(vertical='center', horizontal='center')
    # sheet.font = Font(name=u'微軟正黑', size=10)
    # sheet.alignment = Alignment(vertical='center', horizontal='center')
    # sheet.merge_cells(start_row=1, start_column=11, end_row=1+len(datas), end_column=11)
    workbook.save('test2.xlsx')

# open_and_login(driver)

try:
    element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'notification-list')))
except Exception as e:
    print("找不到元素notification-list")

# element = driver.find_elements_by_class_name('notification-list')
#print(element)
# date, account_id, name, delivery, shop, itmes, color, size,  total = [], [], [], [], [], [], [], [], []
print("要爬多少頁:")
num_of_pages = int(input())
current_page = 1
for i in range(num_of_pages-1):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    next_page_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CLASS_NAME, 'eds-button.eds-button--small.eds-button--frameless.eds-button--block.eds-pager__button-next'))
    )
    next_page_button.click()
    current_page += 1


while True:
    time.sleep(0.5)

    notification_items = driver.find_elements(By.CLASS_NAME, 'notification-item')
    t=0
    for item in reversed(notification_items):
    # for item in notification_items:
        t = t + 1
        if t==5:
            driver.execute_script("window.scrollTo(0, 0);")
        print("into loop i:{}".format(t))
        #badge_sup = item.find_elements(By.CSS_SELECTOR, '.shopee-badge-x__sup shopee-badge-x__sup--dot.shopee-badge-x__sup--fixed')
        #if badge_sup and badge_sup[0].is_displayed(): #代表是新的
        # time.sleep(2)
        try:
            item_button = WebDriverWait(item, 10).until(
                EC.element_to_be_clickable((By.CLASS_NAME, 'container-item.clickable'))
            )
            if "你有一筆貨到付款的訂單" in item_button.find_element_by_class_name('title').text:
                ActionChains(driver).move_to_element(item_button).perform()
                item_button.click()
                print("點開頁面")
                # time.sleep(2)
                driver.switch_to.window(driver.window_handles[1])
                # print('切換成功')
                # date = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME,"time")))
                description_element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, 'description'))
                )

                # 再次使用WebDriverWait等待time元素在description元素內部可見
                # 注意：此處假設description元素內部只有一個time類的元素
                date= WebDriverWait(description_element, 10).until(
                    EC.visibility_of_element_located((By.CLASS_NAME, 'time'))
                )

                date = transition_time(date.text)
                # print('date success')
                # print(driver.find_element_by_class_name('time').text) 
                account_id = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "username.text-overflow"))).text
                # print('account_id success')
                # print(driver.find_element_by_class_name('username.text-overflow').text) #account id
                name = ""
                # all_cards = WebDriverWait(driver, 10).until(
                #     EC.presence_of_all_elements_located((By.CLASS_NAME, "card-style.shopee-card"))
                # )
                # info = driver.find_elements_by_class_name('card-style.shopee-card')
                # info = all_cards[2].text
                # pattern = re.compile(r'買家收件地址\n(.+?)\n')
                # match = pattern.search(info)
                # if match:
                #     buyer_address = match.group(1)
                #     name = buyer_address.split(',')[0]
                delivery_way = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "carrier"))).text
                # delivery_way = driver.find_element_by_class_name('carrier').text
                print('')
                delivery_way = transition_delivery_way(delivery_way)
                delivery_shop = ""
                phone=""
                total = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "amount"))).text
                # total = driver.find_element_by_class_name('amount').text
                total = total[3:] # reove NT$
                total_money = int(total.replace(',', '')) # replace 1,386 to 1386

                # print(driver.find_element_by_class_name('shopee-card__content').find_elements_by_class_name('body').text)
                # product_list = driver.find_element_by_class_name('product-list')
                # qty_list = product_list.find_elements_by_class_name('qty')[1:]
                # products = product_list.find_elements_by_class_name('product-item.product')
                product_list = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "product-list")))
                qty_list =  wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "qty")))[1:]
                products = WebDriverWait(product_list, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, "product-item.product")))
                data=[]
                for product, qty_element in zip(products, qty_list):
                    product_name, color, size = process_product_name(product.text)
                    # Extract qty using Selenium
                    qty_value = int(qty_element.text.strip()) if qty_element else 0
                    # Append data based on qty
                    for _ in range(qty_value):
                        data.append([date, account_id, name, phone, delivery_way, delivery_shop, product_name, color, size,
                                     total_money])

                print(data)
                write_order_xlsx(data)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                print('scratch over')
                # time.sleep(1)
        except NoSuchElementException:
            print('Element not found within the given time')
            pass
        # time.sleep(1.5)

    if current_page > 1:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        prev_page_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'eds-button.eds-button--small.eds-button--frameless.eds-button--block.eds-pager__button-prev'))
        )
        # 获取元素的class name
        class_name = prev_page_button.get_attribute('class')
        prev_page_button.click()
        current_page -= 1
    else:
        break

    # time.sleep(3)

print("test over")
driver.quit()


'''
打包執行檔指令:
pyinstaller -F -c --hidden-import=openpyxl.cell._writer Order.py
'''
